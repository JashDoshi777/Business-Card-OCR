import cv2
import numpy as np
import easyocr
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests # You'll need to install this: pip install requests

# --- Configuration ---
input_folder = 'input'
output_folder = 'output'

os.makedirs(input_folder, exist_ok=True)
os.makedirs(output_folder, exist_ok=True)

# --- IMPORTANT: Configure your Search API Key and CX here ---
GOOGLE_SEARCH_API_KEY = "" # Replace with your actual API key
GOOGLE_SEARCH_CX = ""   # Replace with your actual CX ID

# --- 1. OCR Processing Function ---
def perform_ocr(image_path, output_dir, reader_instance):
    """
    Performs OCR on an image and returns the full extracted text.
    Also saves intermediate image processing steps.
    """
    print(f"Starting OCR processing for {image_path}...")
    img = cv2.imread(image_path)
    if img is None:
        print(f"⚠️ Could not find or read the file: {image_path}")
        return None

    base_filename = os.path.basename(image_path).split('.')[0]

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_1_gray.jpg'), gray)

    adaptive = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 15, 9
    )
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_2_adaptive.jpg'), adaptive)

    inverted = cv2.bitwise_not(adaptive)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_3_inverted.jpg'), inverted)

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (70, 1))
    detected_lines = cv2.morphologyEx(inverted, cv2.MORPH_OPEN, kernel, iterations=3)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_4_detected_lines.jpg'), detected_lines)

    no_lines = cv2.subtract(inverted, detected_lines)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_5_no_lines.jpg'), no_lines)

    cleaned = cv2.bitwise_not(no_lines)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_6_cleaned_final.jpg'), cleaned)

    blurred = cv2.GaussianBlur(cleaned, (0,0), 3)
    sharpened = cv2.addWeighted(cleaned, 1.5, blurred, -0.5, 0)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_7_sharpened.jpg'), sharpened)

    scale = 2
    upscaled = cv2.resize(sharpened, (0,0), fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    cv2.imwrite(os.path.join(output_dir, f'{base_filename}_8_upscaled.jpg'), upscaled)

    result = reader_instance.readtext(upscaled)

    ocr_text_lines = [text for (bbox, text, conf) in result]
    full_ocr_text = "\n".join(ocr_text_lines)

    text_output_path = os.path.join(output_dir, f'{base_filename}_final_easyocr_output.txt')
    with open(text_output_path, 'w', encoding='utf-8') as f:
        f.write(full_ocr_text)

    print(f"✅ OCR Done for {image_path}! Raw output saved to: {text_output_path}")
    return full_ocr_text

# --- Helper Function for Web Search (Local Implementation) ---
def _perform_google_search(query):
    """
    Performs a Google Custom Search API query.
    Requires GOOGLE_SEARCH_API_KEY and GOOGLE_SEARCH_CX to be set.
    """
    if not GOOGLE_SEARCH_API_KEY or not GOOGLE_SEARCH_CX:
        print("❌ Google Search API Key or CX not configured. Skipping web search.")
        return []

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": GOOGLE_SEARCH_API_KEY,
        "cx": GOOGLE_SEARCH_CX,
        "q": query,
        "num": 5 # Number of results to fetch
    }
    try:
        response = requests.get(url, params=params)
        response.raise_for_status() # Raise an exception for HTTP errors
        return response.json().get('items', [])
    except requests.exceptions.RequestException as e:
        print(f"❌ Error during Google Search API call: {e}")
        return []

# --- Company Name Correction using Local Web Search ---
def correct_company_name_with_local_search(current_company_name, owner_name_list, email, address):
    """
    Attempts to correct a company name using email domain and local web search (Google Custom Search API).
    Note: owner_name_list is now a list.
    """
    corrected_name = current_company_name
    owner_name = owner_name_list[0] if owner_name_list else None # Use the first owner name for search

    # --- 1. Prioritize Correction from Email Domain ---
    if email:
        try:
            domain_match = re.search(r'@([a-zA-Z0-9.-]+)\.', email)
            if domain_match:
                domain_part = domain_match.group(1)
                cleaned_domain = re.sub(r'\.(com|org|net|co\.uk|in|io|biz|info|us|ca|gov|edu|me|app)$', '', domain_part, flags=re.IGNORECASE)
                cleaned_domain = cleaned_domain.replace('-', ' ').replace('.', ' ').strip()
                suggested_from_email = ' '.join(word.capitalize() for word in cleaned_domain.split())

                if suggested_from_email and (
                    not current_company_name or
                    len(current_company_name.split()) < 3 or
                    len(set(current_company_name.lower().split()) & set(suggested_from_email.lower().split())) < 1
                ):
                    print(f"💡 Suggesting company name from email domain: '{suggested_from_email}'")
                    corrected_name = suggested_from_email
                    return corrected_name
        except Exception as e:
            print(f"Error processing email domain for correction: {e}")

    # --- 2. Web Search Fallback ---
    is_suspicious = not corrected_name or \
                    len(corrected_name.split()) < 2 or \
                    re.search(r'\d', corrected_name)

    if is_suspicious:
        print(f"🌐 Attempting web search for company name correction for: '{current_company_name}'")
        search_queries = []
        if current_company_name:
            search_queries.append(f'"{current_company_name}" official website')
        if owner_name and current_company_name:
            search_queries.append(f'"{owner_name}" "{current_company_name}"')
        if email:
            search_queries.append(f'"{email}" company')
            if 'suggested_from_email' in locals() and suggested_from_email:
                search_queries.append(f'"{suggested_from_email}" official website')
        if address and current_company_name:
            search_queries.append(f'"{current_company_name}" "{address}"')

        queries_to_execute = search_queries[:2] # Limit to top 2 relevant queries

        for query in queries_to_execute:
            search_results = _perform_google_search(query)
            for res in search_results:
                snippet_lower = res.get('snippet', '').lower()
                title_lower = res.get('title', '').lower()

                if email and 'suggested_from_email' in locals() and suggested_from_email.lower() in snippet_lower:
                    if suggested_from_email.lower() in title_lower:
                        print(f"✅ Found strong match in search title: '{res.get('title')}'")
                        return res.get('title').split(' - ')[0].strip()
                    elif suggested_from_email.lower() in snippet_lower:
                        print(f"✅ Found strong match in search snippet for email domain.")
                        return suggested_from_email

                if current_company_name and current_company_name.lower() in title_lower:
                    print(f"✅ Found current company name in search title: '{res.get('title')}'")
                    if len(res.get('title')) > len(current_company_name) + 5:
                        return res.get('title').split(' - ')[0].strip()

                potential_company_matches = re.findall(r'\b[A-Z][a-z]+(?: [A-Z][a-z]+)*(?: Inc\.| Ltd\.| Corp\.| Group| Solutions| Tech| Industries| Co\.| Pvt\.| LLP| PLC| Ventures| Enterprises)\b', snippet_lower, re.IGNORECASE)
                if potential_company_matches:
                    best_match = max(potential_company_matches, key=len)
                    print(f"🔍 Found potential company name in snippet: '{best_match}'")
                    if not corrected_name or (len(best_match) > len(corrected_name) and best_match.lower() != corrected_name.lower()):
                        corrected_name = best_match.strip()
                        return corrected_name
        
    return corrected_name

# --- 2. Generalized Function to Parse Extracted Text ---
def extract_business_card_info_generalized(ocr_text):
    """
    Extracts structured information (Company, Owner, Email, Number, Address)
    from raw OCR text using generalized heuristics.
    This version now handles multiple owner names and phone numbers,
    and fixes the address extraction issue.
    """
    info = {
        "Company Name": None,
        "Owner Name": [],
        "Email": None,
        "Number": [],
        "Address": [], # Initialize as an empty list to consistently collect lines
    }

    lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]

    email_pattern = r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b'
    phone_pattern = r'(?:\+\d{1,3}[-.\s]?)?(?:\(\d{1,5}\)[-.\s]?)?\d{1,5}[-.\s]?\d{1,5}[-.\s]?\d{1,5}[-.\s]?\d{1,5}'


    found_numbers = set()
    for i, line in enumerate(lines):
        if not line: continue

        emails = re.findall(email_pattern, line, re.IGNORECASE)
        if emails and not info["Email"]:
            info["Email"] = emails[0]

        matches = re.finditer(phone_pattern, line)
        for match in matches:
            potential_number = match.group(0)
            cleaned_number = re.sub(r'[^\d+]', '', potential_number)
            if len(cleaned_number) >= 7 and cleaned_number not in found_numbers:
                info["Number"].append(cleaned_number)
                found_numbers.add(cleaned_number)


    filtered_lines = [line for line in lines if not (
        (info["Email"] and info["Email"].lower() in line.lower()) or
        any(num in re.sub(r'[^\d+]', '', line) for num in info["Number"])
    )]

    job_titles = ["ceo", "cto", "cfo", "md", "director", "manager", "president", "founder", "owner", "partner", "head of", "vp", "chief"]
    company_indicators = ["inc", "llc", "ltd", "corp", "group", "solutions", "tech", "industries", "company", "co.", "pvt", "gmbh", "s.a.", "llp", "plc", "associates", "ventures", "enterprises"]
    address_indicators = ["street", "road", "avenue", "st", "rd", "ave", "blvd", "lane", "ln", "suite", "apt", "floor", "building", "centre", "center", "point", "city", "state", "zip", "postcode", "p.o. box", "po box", "india", "usa", "uk", "mumbai", "delhi", "london", "new york", "dist.", "county", "prov.", "gpo"]
    postal_code_pattern = r'\b\d{5}(-\d{4})?\b|\b[A-Za-z]\d[A-Za-z][ -]?\d[A-Za-z]\d\b|\b\d{6}\b'

    found_names = set()
    seen_address_parts = set() # Initialize set for address parts here

    for i, line in enumerate(filtered_lines):
        line_lower = line.lower()

        words = line.split()
        if len(words) >= 2 and all(word[0].isupper() for word in words[:2]) and \
           not any(indicator in line_lower for indicator in company_indicators + address_indicators) and \
           not re.search(r'\d', line) and \
           not any(title in line_lower for title in job_titles):
            if line not in found_names:
                info["Owner Name"].append(line)
                found_names.add(line)
        
        if any(title in line_lower for title in job_titles):
            if i > 0:
                prev_line = filtered_lines[i-1]
                if len(prev_line.split()) >= 2 and all(word[0].isupper() for word in prev_line.split()[:2]) and \
                   not any(indicator in prev_line.lower() for indicator in company_indicators + address_indicators + job_titles) and \
                   not re.search(r'\d', prev_line):
                    if prev_line not in found_names:
                        info["Owner Name"].append(prev_line.strip())
                        found_names.add(prev_line)

        if any(indicator in line_lower for indicator in company_indicators) or \
           (line.isupper() and len(words) > 1 and len(line) > 5) or \
           (len(words) > 2 and all(word[0].isupper() for word in words) and not re.search(r'\d', line) and \
            not any(title in line_lower for title in job_titles)):
            if not info["Company Name"]:
                info["Company Name"] = line.strip()

        # Check for Address (collect all plausible address lines)
        if re.search(r'\d', line) or any(indicator in line_lower for indicator in address_indicators) or \
           re.search(postal_code_pattern, line):
            cleaned_line = line.replace(";", ",").replace(":", ",").strip()
            if cleaned_line and cleaned_line not in seen_address_parts:
                info["Address"].append(cleaned_line)
                seen_address_parts.add(cleaned_line)

    # --- Post-processing and Formatting for Excel ---
    if info["Owner Name"]:
        info["Owner Name"] = ", ".join(sorted(list(set(info["Owner Name"]))))
    else:
        info["Owner Name"] = None

    if info["Number"]:
        info["Number"] = ", ".join(sorted(list(set(info["Number"]))))
    else:
        info["Number"] = None

    if info["Address"]:
        # Join the collected address lines directly
        info["Address"] = ", ".join(info["Address"])
        info["Address"] = re.sub(r',+', ',', info["Address"]).strip()
        info["Address"] = re.sub(r'\s+,', ',', info["Address"]).strip()
        info["Address"] = re.sub(r',\s+', ', ', info["Address"]).strip()
    else:
        info["Address"] = None

    # --- Company Name Correction using Local Web Search ---
    try:
        corrected_company_name = correct_company_name_with_local_search(
            info["Company Name"],
            info["Owner Name"].split(', ') if info["Owner Name"] else [],
            info["Email"],
            info["Address"] # Pass the now-string address
        )
        if corrected_company_name:
            info["Company Name"] = corrected_company_name
    except Exception as e:
        print(f"❌ Error during company name web correction: {e}")

    return info

# --- 3. Function to Create and Populate Excel File ---
def create_excel_sheet(data_list, filename="business_card_data.xlsx"):
    """
    Creates an Excel workbook and populates it with business card data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Card Contacts"

    headers = ["Company Name", "Owner Name", "Email", "Number", "Address"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[chr(64 + col_idx)].width = 25

    for row_data in data_list:
        row_values = [row_data.get(header, "") for header in headers]
        ws.append(row_values)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    try:
        wb.save(filename)
        print(f"✅ Successfully created '{filename}' in the '{output_folder}' directory.")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")

# --- Main Execution Flow ---
if __name__ == "__main__":
    all_business_cards_data = []
    
    print("Initializing EasyOCR reader...")
    reader = easyocr.Reader(['en'], gpu=False)
    print("EasyOCR reader initialized.")

    image_files = [f for f in os.listdir(input_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]

    if not image_files:
        print(f"No image files found in the '{input_folder}' directory. Please place your business card images there.")
    else:
        for image_file in image_files:
            image_path = os.path.join(input_folder, image_file)
            
            full_ocr_text = perform_ocr(image_path, output_folder, reader)
            
            if full_ocr_text:
                print(f"\n--- Extracted Text from {image_file} for Parsing ---")
                print(full_ocr_text)
                print("----------------------------------------------------\n")

                print(f"Starting text parsing and structuring for {image_file}...")
                extracted_info = extract_business_card_info_generalized(full_ocr_text)
                all_business_cards_data.append(extracted_info)

                print("\nParsed Information:")
                for key, value in extracted_info.items():
                    print(f"{key}: {value}")
                print("--------------------------")
            else:
                print(f"Skipping parsing for {image_file} due to OCR failure.")

        output_excel_filename = os.path.join(output_folder, "business_card_contacts_from_images.xlsx")
        create_excel_sheet(all_business_cards_data, output_excel_filename)

    print("\nScript finished.")
